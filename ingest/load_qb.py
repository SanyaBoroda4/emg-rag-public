"""Load the QuickBooks invoice export (XLSX) into invoices and invoice_lines.

Source: /opt/emg-rag/raw/qb/EMG_All_Invoices_2026-07-29.xlsx, sheet "Invoices",
one row per invoice. The "Line Items" column is a display summary of the form
    Service:Labor x23 $810.75 | Materials:Material, ... x1 $600.00 | ...
It is parsed best-effort into invoice_lines (each segment matched from the
right: " x<qty> $<amount>" suffix, everything before it is the item name;
unit_price is derived as amount/qty). The verbatim string is stored in
invoices.line_summary_raw (added by sql/002).

Idempotent: invoices upsert on doc_number; invoice_lines (identity PK) are
delete-then-insert per doc_number. Running twice yields identical counts.
"""

import re
import sys
import traceback
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from ingest.db import get_conn

XLSX_PATH = Path("/opt/emg-rag/raw/qb/EMG_All_Invoices_2026-07-29.xlsx")

# " x<qty> $<amount>" at the end of a segment; name is everything before it.
LINE_RE = re.compile(r"^(?P<name>.*) x(?P<qty>[\d.]+) \$(?P<amt>-?[\d,]+\.?\d*)$")
# Some segments carry no qty at all ("Additional:Sales $0.00") — qty stays NULL.
LINE_NO_QTY_RE = re.compile(r"^(?P<name>.*) \$(?P<amt>-?[\d,]+\.?\d*)$")

INVOICE_SQL = """
    INSERT INTO invoices (doc_number, qb_id, customer_name, txn_date, due_date,
                          total_amt, balance, status, customer_memo, private_note,
                          bill_email, qb_created, qb_updated, line_summary_raw)
    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    ON CONFLICT (doc_number) DO UPDATE SET
        qb_id = EXCLUDED.qb_id, customer_name = EXCLUDED.customer_name,
        txn_date = EXCLUDED.txn_date, due_date = EXCLUDED.due_date,
        total_amt = EXCLUDED.total_amt, balance = EXCLUDED.balance,
        status = EXCLUDED.status, customer_memo = EXCLUDED.customer_memo,
        private_note = EXCLUDED.private_note, bill_email = EXCLUDED.bill_email,
        qb_created = EXCLUDED.qb_created, qb_updated = EXCLUDED.qb_updated,
        line_summary_raw = EXCLUDED.line_summary_raw
"""


def as_date(v):
    """Cells arrive as 'YYYY-MM-DD' strings (or None/''); pass through as date."""
    if v is None:
        return None
    s = str(v).strip()
    return s[:10] if s else None


def parse_lines(doc_number: str, summary):
    """Best-effort parse of the Line Items display string. Returns (rows, fails)."""
    rows, fails = [], 0
    if not summary:
        return rows, fails
    for i, seg in enumerate(str(summary).split(" | "), start=1):
        seg = seg.strip()
        m = LINE_RE.match(seg)
        if m:
            qty = float(m.group("qty"))
        else:
            m = LINE_NO_QTY_RE.match(seg)
            qty = None
        if not m:
            # remaining failures are summaries QB truncated with a trailing "…"
            fails += 1
            continue
        amount = float(m.group("amt").replace(",", ""))
        unit = round(amount / qty, 4) if qty else None
        rows.append((doc_number, i, m.group("name"), None, qty, unit, amount))
    return rows, fails


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"FAIL: {XLSX_PATH} not found — stopping, will not re-pull from "
              f"QuickBooks", file=sys.stderr)
        return 1

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO pipeline_runs (source, status) "
                "VALUES ('qb', 'running') RETURNING run_id")
            run_id = cur.fetchone()[0]
        conn.commit()

        try:
            wb = load_workbook(XLSX_PATH, read_only=True)
            ws = wb["Invoices"]
            n_inv = n_lines = seg_fails = 0
            inv_batch, line_batch, doc_batch = [], [], []

            def flush(cur):
                cur.executemany(INVOICE_SQL, inv_batch)
                cur.execute("DELETE FROM invoice_lines WHERE doc_number = ANY(%s)",
                            (doc_batch,))
                if line_batch:
                    cur.executemany(
                        "INSERT INTO invoice_lines (doc_number, line_num, "
                        "item_name, description, qty, unit_price, amount) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s)", line_batch)

            for row in ws.iter_rows(min_row=2, values_only=True):
                (doc, qb_id, customer, inv_date, due_date, total, balance,
                 status, memo, note, email, line_items, created, updated) = row
                if doc is None:
                    continue
                doc = str(doc).strip()
                inv_batch.append((
                    doc, str(qb_id) if qb_id is not None else None, customer,
                    as_date(inv_date), as_date(due_date), total, balance, status,
                    memo, note, email, as_date(created), as_date(updated),
                    str(line_items) if line_items is not None else None))
                doc_batch.append(doc)
                rows, fails = parse_lines(doc, line_items)
                line_batch.extend(rows)
                seg_fails += fails
                n_inv += 1
                n_lines += len(rows)
                if len(inv_batch) >= 500:
                    with conn.cursor() as cur:
                        flush(cur)
                    conn.commit()
                    inv_batch, line_batch, doc_batch = [], [], []
            if inv_batch:
                with conn.cursor() as cur:
                    flush(cur)
                conn.commit()
            wb.close()

            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE pipeline_runs SET finished_at = now(), status = 'ok', "
                    "rows_in = %s, rows_out = %s WHERE run_id = %s",
                    (n_inv, n_inv + n_lines, run_id))
            conn.commit()
            print(f"invoices: {n_inv} | invoice_lines: {n_lines} | "
                  f"unparseable line segments: {seg_fails}")
        except Exception:
            conn.rollback()
            err = traceback.format_exc()
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE pipeline_runs SET finished_at = now(), "
                    "status = 'failed', error = %s WHERE run_id = %s",
                    (err, run_id))
            conn.commit()
            print(err, file=sys.stderr)
            return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
